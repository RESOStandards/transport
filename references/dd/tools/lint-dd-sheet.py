#!/usr/bin/env python3
"""
Lint and normalize a DD XLSX reference sheet.

Uses openpyxl (not SheetJS / xlsx) because openpyxl preserves cell formatting,
comments, merged cells, column widths, and file size on round-trip. SheetJS
strips formatting on write and roughly doubles file size — do not switch back.

Rewrites the dd.reso.org URLs deterministically on every run, on any DD
version. Six per-row targets (cell value and/or hyperlink, all pointing to
the same canonical URL):

  Fields tab
    - ResourceName cell:  hyperlink         -> /DD{ver}/{ResourceName}/
    - StandardName cell:  hyperlink         -> /DD{ver}/{ResourceName}/{StandardName}/
    - WikiPageUrl cell:   value + hyperlink -> /DD{ver}/{ResourceName}/{StandardName}/

  Lookups tab
    - LookupName cell:           hyperlink         -> /DD{ver}/lookups/{LookupName}/
    - StandardLookupValue cell:  hyperlink         -> /DD{ver}/lookups/{LookupName}/{enc(StandardLookupValue)}/
    - WikiPageUrl cell:          value + hyperlink -> /DD{ver}/lookups/{LookupName}/{enc(StandardLookupValue)}/

The lookup-value URL uses StandardLookupValue (display name, URL-encoded) — that
is what the live dd.reso.org routes match. WikiPageTitle and LegacyODataValue
columns are left untouched.

Usage: python3 lint-dd-sheet.py <input.xlsx> <version> [output.xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook
from openpyxl.worksheet.hyperlink import Hyperlink


# Match JavaScript's encodeURIComponent: encode everything except
# A-Z a-z 0-9 - _ . ~ ! * ' ( )
def enc(s: str) -> str:
    return quote(str(s), safe="!~*'()")


def header_indexes(ws) -> dict[str, int]:
    return {
        str(cell.value): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def set_hyperlink(cell, target: str) -> None:
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=target)


def set_value_and_hyperlink(cell, value: str, target: str) -> None:
    cell.value = value
    cell.hyperlink = Hyperlink(ref=cell.coordinate, target=target)


# RESO DD enumeration data types — only these legitimately carry a LookupStatus.
ENUM_DATA_TYPES = {"String List, Single", "String List, Multi"}


def trim_whitespace(ws, stats: dict) -> None:
    """Strip leading/trailing whitespace from every text cell and hyperlink target — a deterministic
    cleanup like the URL canonicalization. DD text (definitions, names, values) should never carry
    surrounding whitespace; it silently corrupts downstream matching and display."""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v != v.strip():
                cell.value = v.strip()
                stats["whitespace_trimmed"] += 1
            hl = getattr(cell, "hyperlink", None)
            target = getattr(hl, "target", None)
            if isinstance(target, str) and target != target.strip():
                hl.target = target.strip()


def check_type_status_agreement(ws) -> list[str]:
    """SimpleDataType and LookupStatus must agree: a field is an enumeration IF AND ONLY IF its data
    type is 'String List, Single/Multi'. Flag both directions of disagreement —
      (a) a LookupStatus on a non-enumeration type — e.g. BuiltPre1978YN (a Boolean carrying 'Open',
          fixed by clearing the status) or HistoryTransactional.ClassName (a String that should be
          'String List, Single', fixed by correcting the type); and
      (b) an enumeration data type carrying no LookupStatus.
    Warn rather than auto-fix — which side is authoritative (clear the status vs. correct the type)
    depends on the field's true nature and is a human call. Warnings are qualified by Resource.Field
    because a field name (e.g. ClassName) can appear on many resources with differing rows."""
    headers = header_indexes(ws)
    c_type = headers.get("SimpleDataType")
    c_ls = headers.get("LookupStatus")
    c_field = headers.get("StandardName")
    c_res = headers.get("ResourceName")
    if not (c_type and c_ls and c_field):
        return []
    warnings: list[str] = []
    for row in ws.iter_rows(min_row=2):
        field_name = row[c_field - 1].value
        if not field_name:
            continue
        data_type = row[c_type - 1].value
        lookup_status = row[c_ls - 1].value
        resource = row[c_res - 1].value if c_res else None
        label = f"{resource}.{field_name}" if resource else field_name
        is_enum_type = str(data_type) in ENUM_DATA_TYPES
        has_status = lookup_status not in (None, "")
        if has_status and not is_enum_type:
            warnings.append(
                f"{label}: SimpleDataType={data_type!r} carries LookupStatus={lookup_status!r} "
                f"but is not an enumeration data type"
            )
        elif is_enum_type and not has_status:
            warnings.append(
                f"{label}: SimpleDataType={data_type!r} is an enumeration data type "
                f"but carries no LookupStatus"
            )
    return warnings


def lint_fields(ws, base: str, stats: dict) -> None:
    headers = header_indexes(ws)
    required = ("ResourceName", "StandardName", "WikiPageUrl")
    if not all(c in headers for c in required):
        raise SystemExit(f"Fields tab missing one of: {required}")

    c_resource = headers["ResourceName"]
    c_field = headers["StandardName"]
    c_url = headers["WikiPageUrl"]

    for row in ws.iter_rows(min_row=2):
        stats["rows"] += 1
        resource_cell = row[c_resource - 1]
        field_cell = row[c_field - 1]
        url_cell = row[c_url - 1]

        resource_name = resource_cell.value
        field_name = field_cell.value
        if not resource_name or not field_name:
            stats["skipped"] += 1
            continue

        resource_url = f"{base}/{enc(resource_name)}/"
        field_url = f"{base}/{enc(resource_name)}/{enc(field_name)}/"

        if not resource_cell.hyperlink or resource_cell.hyperlink.target != resource_url:
            stats["resource_link"] += 1
        set_hyperlink(resource_cell, resource_url)

        if not field_cell.hyperlink or field_cell.hyperlink.target != field_url:
            stats["field_link"] += 1
        set_hyperlink(field_cell, field_url)

        if url_cell.value != field_url:
            stats["wiki_value"] += 1
        if not url_cell.hyperlink or url_cell.hyperlink.target != field_url:
            stats["wiki_link"] += 1
        set_value_and_hyperlink(url_cell, field_url, field_url)


def lint_lookups(ws, base: str, stats: dict) -> None:
    headers = header_indexes(ws)
    value_col_name = (
        "StandardLookupValue" if "StandardLookupValue" in headers
        else "LookupDisplayName" if "LookupDisplayName" in headers
        else None
    )
    required = ("LookupName", "WikiPageUrl")
    if value_col_name is None or not all(c in headers for c in required):
        raise SystemExit(
            "Lookups tab missing one of: LookupName, "
            "StandardLookupValue (or LookupDisplayName), WikiPageUrl"
        )

    c_name = headers["LookupName"]
    c_value = headers[value_col_name]
    c_url = headers["WikiPageUrl"]

    for row in ws.iter_rows(min_row=2):
        stats["rows"] += 1
        name_cell = row[c_name - 1]
        value_cell = row[c_value - 1]
        url_cell = row[c_url - 1]

        lookup_name = name_cell.value
        standard_value = value_cell.value
        if not lookup_name or not standard_value:
            stats["skipped"] += 1
            continue

        enum_url = f"{base}/lookups/{enc(lookup_name)}/"
        value_url = f"{base}/lookups/{enc(lookup_name)}/{enc(standard_value)}/"

        if not name_cell.hyperlink or name_cell.hyperlink.target != enum_url:
            stats["name_link"] += 1
        set_hyperlink(name_cell, enum_url)

        if not value_cell.hyperlink or value_cell.hyperlink.target != value_url:
            stats["value_link"] += 1
        set_hyperlink(value_cell, value_url)

        if url_cell.value != value_url:
            stats["wiki_value"] += 1
        if not url_cell.hyperlink or url_cell.hyperlink.target != value_url:
            stats["wiki_link"] += 1
        set_value_and_hyperlink(url_cell, value_url, value_url)


def main() -> None:
    args = sys.argv[1:]
    if len(args) < 2:
        print("Usage: lint-dd-sheet.py <input.xlsx> <version> [output.xlsx]")
        sys.exit(0)

    input_path = Path(args[0]).resolve()
    version = args[1]
    output_path = Path(args[2]).resolve() if len(args) >= 3 else input_path

    base = f"https://dd.reso.org/DD{version}"
    wb = load_workbook(input_path)

    field_stats = {
        "rows": 0, "resource_link": 0, "field_link": 0,
        "wiki_value": 0, "wiki_link": 0, "skipped": 0, "whitespace_trimmed": 0,
    }
    lookup_stats = {
        "rows": 0, "name_link": 0, "value_link": 0,
        "wiki_value": 0, "wiki_link": 0, "skipped": 0, "whitespace_trimmed": 0,
    }

    type_status_warnings: list[str] = []
    if "Fields" in wb.sheetnames:
        trim_whitespace(wb["Fields"], field_stats)
        lint_fields(wb["Fields"], base, field_stats)
        type_status_warnings = check_type_status_agreement(wb["Fields"])
    if "Lookups" in wb.sheetnames:
        trim_whitespace(wb["Lookups"], lookup_stats)
        lint_lookups(wb["Lookups"], base, lookup_stats)

    wb.save(output_path)

    print(f"Linted: {output_path}")
    print(
        f"  Fields:  rows={field_stats['rows']} "
        f"resource-link={field_stats['resource_link']} "
        f"field-link={field_stats['field_link']} "
        f"wiki-value={field_stats['wiki_value']} "
        f"wiki-link={field_stats['wiki_link']} "
        f"skipped={field_stats['skipped']} "
        f"whitespace-trimmed={field_stats['whitespace_trimmed']}"
    )
    print(
        f"  Lookups: rows={lookup_stats['rows']} "
        f"name-link={lookup_stats['name_link']} "
        f"value-link={lookup_stats['value_link']} "
        f"wiki-value={lookup_stats['wiki_value']} "
        f"wiki-link={lookup_stats['wiki_link']} "
        f"skipped={lookup_stats['skipped']} "
        f"whitespace-trimmed={lookup_stats['whitespace_trimmed']}"
    )

    if type_status_warnings:
        print(
            f"\n  WARNING: {len(type_status_warnings)} field(s) have a SimpleDataType / LookupStatus "
            f"disagreement (review — clear the status or correct the data type):"
        )
        for warning in type_status_warnings:
            print(f"    - {warning}")


if __name__ == "__main__":
    main()
