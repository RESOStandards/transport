"""
Tests for lint-dd-sheet.py — the DD XLSX linter.

The linter is cert-adjacent tooling (it normalizes the source-of-truth DD sheet), so its checks must
hold 100%. These tests exercise the pure check/normalize functions against in-memory workbooks built
with openpyxl — no fixture files needed.

The linter is a hyphenated script (not an importable module name), so it is loaded by path. Run with
the DD-lint venv that already has openpyxl + pytest (see the "DD reference tooling" section of
CLAUDE.md):

    references/dd/tools/.venv-dd-lint/bin/python -m pytest references/dd/tools/test_lint_dd_sheet.py

The venv is intentionally kept out of git. Run these when touching the linter or refreshing a DD sheet.
"""

from __future__ import annotations

import importlib.util
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink

# Load the hyphenated script by path.
_spec = importlib.util.spec_from_file_location(
    "lint_dd_sheet", Path(__file__).parent / "lint-dd-sheet.py"
)
assert _spec and _spec.loader
lint = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(lint)

BASE = "https://dd.reso.org/DD2.1"


# --- enc: encodeURIComponent parity --------------------------------------------------------------


def test_enc_matches_encodeuricomponent():
    # Spaces and most punctuation are percent-encoded; the JS-unreserved set is preserved.
    assert lint.enc("Public Sewer") == "Public%20Sewer"
    assert lint.enc("Appraiser ") == "Appraiser%20"  # trailing space encodes, not trimmed
    assert lint.enc("A&B") == "A%26B"
    assert lint.enc("R/W") == "R%2FW"
    assert lint.enc("a-b_c.d~e!f*g'h(i)") == "a-b_c.d~e!f*g'h(i)"  # unreserved, untouched


# --- trim_whitespace -----------------------------------------------------------------------------


def _make_sheet(rows, title="Fields", header=None):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    ws.append(header or ["col"])
    for r in rows:
        ws.append(list(r))
    return ws


def test_trim_whitespace_trims_values_and_counts():
    ws = _make_sheet([["  hello  "], ["world "], ["clean"], ["   "]])
    stats = {"whitespace_trimmed": 0}
    lint.trim_whitespace(ws, stats)
    # row indexing: header is row 1, data starts row 2.
    assert ws.cell(row=2, column=1).value == "hello"
    assert ws.cell(row=3, column=1).value == "world"
    assert ws.cell(row=4, column=1).value == "clean"  # unchanged, not counted
    assert ws.cell(row=5, column=1).value == ""  # all-whitespace normalizes to empty
    assert stats["whitespace_trimmed"] == 3  # rows 2, 3, 5 (not the already-clean row 4)


def test_trim_whitespace_trims_hyperlink_targets():
    ws = _make_sheet([["value"]])
    cell = ws.cell(row=2, column=1)
    cell.hyperlink = Hyperlink(ref="A2", target="  https://dd.reso.org/x  ")
    stats = {"whitespace_trimmed": 0}
    lint.trim_whitespace(ws, stats)
    assert cell.hyperlink.target == "https://dd.reso.org/x"


def test_trim_whitespace_idempotent():
    ws = _make_sheet([["  hello  "]])
    stats = {"whitespace_trimmed": 0}
    lint.trim_whitespace(ws, stats)
    first = stats["whitespace_trimmed"]
    lint.trim_whitespace(ws, stats)  # second pass
    assert stats["whitespace_trimmed"] == first  # nothing more to trim


# --- check_type_status_agreement -----------------------------------------------------------------

_FIELDS_HEADER = ["ResourceName", "StandardName", "SimpleDataType", "LookupStatus"]


def _fields_sheet(rows):
    return _make_sheet(rows, title="Fields", header=_FIELDS_HEADER)


def test_agreement_passes_consistent_rows():
    ws = _fields_sheet([
        ["Property", "StandardStatus", "String List, Single", "Locked with Enumerations"],
        ["Property", "Appliances", "String List, Multi", "Open with Enumerations"],
        ["Property", "ListPrice", "Number", None],
        ["Property", "PublicRemarks", "String", None],
        ["Media", "MediaModificationTimestamp", "Timestamp", None],
        ["", "", None, None],  # blank trailing row — skipped
    ])
    assert lint.check_type_status_agreement(ws) == []


def test_agreement_flags_lookupstatus_on_non_enum():
    # Direction (a): the two real DD 2.1 cases.
    ws = _fields_sheet([
        ["HistoryTransactional", "ClassName", "String", "Open with Enumerations"],
        ["Property", "BuiltPre1978YN", "Boolean", "Open"],
    ])
    warnings = lint.check_type_status_agreement(ws)
    assert len(warnings) == 2
    assert any(w.startswith("HistoryTransactional.ClassName:") and "not an enumeration" in w
               for w in warnings)
    assert any(w.startswith("Property.BuiltPre1978YN:") and "not an enumeration" in w
               for w in warnings)


def test_agreement_flags_enum_type_without_status():
    # Direction (b): an enum data type with no LookupStatus.
    ws = _fields_sheet([
        ["Property", "OrphanEnum", "String List, Single", None],
        ["Property", "OrphanMulti", "String List, Multi", ""],  # empty string counts as absent
    ])
    warnings = lint.check_type_status_agreement(ws)
    assert len(warnings) == 2
    assert all("carries no LookupStatus" in w for w in warnings)


def test_agreement_qualifies_label_by_resource():
    # ClassName appears on many resources; only one row disagrees — the label must pinpoint it.
    ws = _fields_sheet([
        ["ContactListings", "ClassName", "String List, Single", "Open with Enumerations"],
        ["HistoryTransactional", "ClassName", "String", "Open with Enumerations"],
    ])
    warnings = lint.check_type_status_agreement(ws)
    assert len(warnings) == 1
    assert warnings[0].startswith("HistoryTransactional.ClassName:")


# --- URL canonicalization ------------------------------------------------------------------------


def test_lint_fields_sets_canonical_urls():
    ws = _make_sheet(
        [["Property", "ListPrice", "stale"]],
        title="Fields",
        header=["ResourceName", "StandardName", "WikiPageUrl"],
    )
    stats = {k: 0 for k in ("rows", "resource_link", "field_link", "wiki_value", "wiki_link", "skipped")}
    lint.lint_fields(ws, BASE, stats)
    assert ws.cell(row=2, column=1).hyperlink.target == f"{BASE}/Property/"
    assert ws.cell(row=2, column=2).hyperlink.target == f"{BASE}/Property/ListPrice/"
    assert ws.cell(row=2, column=3).value == f"{BASE}/Property/ListPrice/"
    assert ws.cell(row=2, column=3).hyperlink.target == f"{BASE}/Property/ListPrice/"


def test_lint_lookups_encodes_value_segment():
    ws = _make_sheet(
        [["Sewer", "Public Sewer", "stale"]],
        title="Lookups",
        header=["LookupName", "StandardLookupValue", "WikiPageUrl"],
    )
    stats = {k: 0 for k in ("rows", "name_link", "value_link", "wiki_value", "wiki_link", "skipped")}
    lint.lint_lookups(ws, BASE, stats)
    assert ws.cell(row=2, column=1).hyperlink.target == f"{BASE}/lookups/Sewer/"
    assert ws.cell(row=2, column=2).hyperlink.target == f"{BASE}/lookups/Sewer/Public%20Sewer/"
    assert ws.cell(row=2, column=3).value == f"{BASE}/lookups/Sewer/Public%20Sewer/"
