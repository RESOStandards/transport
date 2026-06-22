#!/usr/bin/env python3
"""
diff-dd-sheet.py — compare two DD reference XLSX workbooks and emit a
markdown summary of the differences.

Use when reviewing incoming DD-sheet update tickets on
RESOStandards/transport (the workgroup contributor attaches a new XLSX
file; we diff it against the current `references/dd/` golden master
before linting and integrating).

Tabs compared:
  - Fields   (key: ResourceName + StandardName)
  - Lookups  (key: LookupName + StandardLookupValue)

For each row, the script compares every cell value AND the hyperlink
target on cells where hyperlinks carry semantic meaning (ResourceName,
StandardName, WikiPageUrl, LookupName, StandardLookupValue). This
mirrors what `lint-dd-sheet.py` normalises and what
`generate-reference-metadata.js` reads.

Usage
-----
Setup (one-time; reuses the venv from lint-dd-sheet.py):

    cd reso-certification
    python3 -m venv .venv-dd-lint
    .venv-dd-lint/bin/pip install openpyxl

Run:

    .venv-dd-lint/bin/python utils/diff-dd-sheet.py \\
        path/to/baseline.xlsx path/to/new.xlsx > diff.md

Or pipe straight to a GitHub issue/PR comment:

    .venv-dd-lint/bin/python utils/diff-dd-sheet.py \\
        baseline.xlsx new.xlsx | gh issue comment 203 \\
        --repo RESOStandards/transport --body-file -

Options
-------
    --json          Emit machine-readable JSON instead of markdown.
    --max-bytes N   Truncate markdown to N bytes (default 60000, leaves
                    headroom under GitHub's 65,536-char comment cap).
                    The truncation marker tells the user where to find
                    the full diff if they need it.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write(
        "openpyxl is required. From the reso-certification dir:\n"
        "  python3 -m venv .venv-dd-lint\n"
        "  .venv-dd-lint/bin/pip install openpyxl\n"
        "  .venv-dd-lint/bin/python utils/diff-dd-sheet.py ...\n"
    )
    sys.exit(1)


# ── tab configuration ───────────────────────────────────────────────────
#
# Keys identify rows across versions; hyperlink-columns are the cells
# whose URL targets are semantically meaningful and must be diffed in
# addition to the cell's text value.

TAB_CONFIG = {
    "Fields": {
        "key_cols": ["ResourceName", "StandardName"],
        "hyperlink_cols": {"ResourceName", "StandardName", "WikiPageUrl"},
    },
    "Lookups": {
        "key_cols": ["LookupName", "StandardLookupValue"],
        "hyperlink_cols": {"LookupName", "StandardLookupValue", "WikiPageUrl"},
    },
}


def _normalize(v):
    """Normalize a cell value for comparison.

    None and empty-string are treated as equivalent. Other values are
    coerced to string and stripped. Numeric types stay distinguishable
    by their string representation (so 0 != "0" is False after coercion,
    which is fine for our purposes).
    """
    if v is None:
        return ""
    return str(v).strip()


def _load_tab(wb, tab_name, config):
    """Build {row_key: {column: (value, hyperlink_target)}} for one tab.

    Returns (rows_dict, header_list). If the tab is missing, returns
    ({}, []) so the diff just shows zero rows on that side.
    """
    if tab_name not in wb.sheetnames:
        return {}, []

    ws = wb[tab_name]
    rows_iter = ws.iter_rows()
    header_row = next(rows_iter, None)
    if header_row is None:
        return {}, []

    headers = [_normalize(c.value) for c in header_row]
    header_idx = {h: i for i, h in enumerate(headers) if h}

    # Validate key columns exist
    for key_col in config["key_cols"]:
        if key_col not in header_idx:
            sys.stderr.write(
                f"warning: tab '{tab_name}' is missing key column '{key_col}'; "
                f"using whatever rows are present, but key collisions are possible.\n"
            )

    rows = {}
    for row in rows_iter:
        # Build the composite key from key columns
        key_parts = []
        for key_col in config["key_cols"]:
            idx = header_idx.get(key_col)
            if idx is None or idx >= len(row):
                key_parts.append("")
            else:
                key_parts.append(_normalize(row[idx].value))
        # Skip totally blank rows (all key parts empty)
        if all(p == "" for p in key_parts):
            continue
        key = "::".join(key_parts)

        cells = {}
        for col, idx in header_idx.items():
            if idx >= len(row):
                continue
            cell = row[idx]
            value = _normalize(cell.value)
            hyperlink = None
            if col in config["hyperlink_cols"] and cell.hyperlink is not None:
                hyperlink = _normalize(cell.hyperlink.target)
            cells[col] = (value, hyperlink)
        rows[key] = cells

    return rows, headers


def _row_diff(old_cells, new_cells, hyperlink_cols):
    """Return list of (column, old_repr, new_repr) for changed cells.

    A cell counts as changed if either the text value or the hyperlink
    target differs. We display both deltas when both move.
    """
    all_cols = set(old_cells.keys()) | set(new_cells.keys())
    changes = []
    for col in all_cols:
        old_val, old_link = old_cells.get(col, ("", None))
        new_val, new_link = new_cells.get(col, ("", None))
        if old_val == new_val and old_link == new_link:
            continue

        # Compose the change description
        if old_val != new_val and old_link != new_link and col in hyperlink_cols:
            old_repr = f"value={_short(old_val)!r}, link={_short(old_link)!r}"
            new_repr = f"value={_short(new_val)!r}, link={_short(new_link)!r}"
        elif old_val != new_val:
            old_repr = repr(_short(old_val))
            new_repr = repr(_short(new_val))
        else:  # hyperlink-only change
            old_repr = f"link={_short(old_link)!r}"
            new_repr = f"link={_short(new_link)!r}"
        changes.append((col, old_repr, new_repr))
    return changes


def _short(s, limit=200):
    """Truncate long strings for readable diff output."""
    if s is None:
        return ""
    s = str(s)
    if len(s) <= limit:
        return s
    return s[:limit] + " […truncated]"


def _diff_tab(tab_name, baseline_rows, new_rows, config):
    """Compare two tabs of rows and return a structured diff dict."""
    baseline_keys = set(baseline_rows.keys())
    new_keys = set(new_rows.keys())

    added = sorted(new_keys - baseline_keys)
    removed = sorted(baseline_keys - new_keys)

    modified = []
    for key in sorted(baseline_keys & new_keys):
        changes = _row_diff(
            baseline_rows[key], new_rows[key], config["hyperlink_cols"]
        )
        if changes:
            modified.append({"key": key, "changes": changes})

    return {
        "tab": tab_name,
        "baseline_count": len(baseline_rows),
        "new_count": len(new_rows),
        "added": [{"key": k, "row": new_rows[k]} for k in added],
        "removed": [{"key": k, "row": baseline_rows[k]} for k in removed],
        "modified": modified,
    }


def _format_markdown(baseline_path, new_path, tab_diffs):
    """Render the diff as a markdown block suitable for a GH comment."""
    out = []
    out.append("## DD Sheet Diff")
    out.append("")
    out.append(f"- **Baseline**: `{Path(baseline_path).name}`")
    out.append(f"- **New**: `{Path(new_path).name}`")
    out.append("")

    # Summary table
    out.append("### Summary")
    out.append("")
    out.append("| Tab | Baseline rows | New rows | Δ | Added | Removed | Modified |")
    out.append("|---|---:|---:|---:|---:|---:|---:|")
    for d in tab_diffs:
        delta = d["new_count"] - d["baseline_count"]
        delta_str = f"+{delta}" if delta > 0 else str(delta) if delta < 0 else "0"
        out.append(
            f"| {d['tab']} | {d['baseline_count']:,} | {d['new_count']:,} | "
            f"{delta_str} | {len(d['added'])} | {len(d['removed'])} | "
            f"{len(d['modified'])} |"
        )
    out.append("")

    # Per-tab detail
    for d in tab_diffs:
        if not (d["added"] or d["removed"] or d["modified"]):
            continue
        out.append(f"### {d['tab']} tab")
        out.append("")

        if d["added"]:
            out.append(f"**Added ({len(d['added'])} rows)**")
            out.append("")
            for entry in d["added"]:
                out.append(f"- `{entry['key']}`")
            out.append("")

        if d["removed"]:
            out.append(f"**Removed ({len(d['removed'])} rows)**")
            out.append("")
            for entry in d["removed"]:
                out.append(f"- `{entry['key']}`")
            out.append("")

        if d["modified"]:
            out.append(f"**Modified ({len(d['modified'])} rows)**")
            out.append("")
            for entry in d["modified"]:
                out.append(f"<details><summary><code>{entry['key']}</code> — {len(entry['changes'])} cell(s) changed</summary>")
                out.append("")
                out.append("| Column | Before | After |")
                out.append("|---|---|---|")
                for col, old, new in entry["changes"]:
                    # Escape pipes in cell contents so markdown table doesn't break
                    old_esc = old.replace("|", "\\|").replace("\n", " ")
                    new_esc = new.replace("|", "\\|").replace("\n", " ")
                    out.append(f"| `{col}` | {old_esc} | {new_esc} |")
                out.append("")
                out.append("</details>")
                out.append("")
        out.append("")

    return "\n".join(out)


def _format_changelog(baseline_path, new_path, tab_diffs, ticket, title, date):
    """Render a compact changelog entry suitable for prepending to CHANGELOG.md.

    Tight summary — per-tab counts and a bulleted list of modified rows
    (without per-cell deltas). The detailed diff goes into the GH
    comment; the changelog is the durable index.
    """
    out = []
    heading = f"## {date}"
    if title:
        heading += f" — {title}"
    out.append(heading)

    if ticket:
        ticket_link = (
            f"**Ticket**: [#{ticket}]("
            f"https://github.com/RESOStandards/transport/issues/{ticket})"
        )
        out.append(ticket_link)

    out.append(f"**Baseline**: `{Path(baseline_path).name}`")
    out.append(f"**New**: `{Path(new_path).name}`")
    out.append("")

    summary_parts = []
    for d in tab_diffs:
        a, r, m = len(d["added"]), len(d["removed"]), len(d["modified"])
        if a or r or m:
            summary_parts.append(f"{d['tab']} {a}/{r}/{m}")
    if summary_parts:
        out.append("**Tab changes** (added / removed / modified): " + ", ".join(summary_parts))
        out.append("")

    for d in tab_diffs:
        if not (d["added"] or d["removed"] or d["modified"]):
            continue
        if d["added"]:
            out.append(f"**{d['tab']} — added ({len(d['added'])})**:")
            for entry in d["added"][:20]:
                out.append(f"- `{entry['key']}`")
            if len(d["added"]) > 20:
                out.append(f"- … {len(d['added']) - 20} more")
            out.append("")
        if d["removed"]:
            out.append(f"**{d['tab']} — removed ({len(d['removed'])})**:")
            for entry in d["removed"][:20]:
                out.append(f"- `{entry['key']}`")
            if len(d["removed"]) > 20:
                out.append(f"- … {len(d['removed']) - 20} more")
            out.append("")
        if d["modified"]:
            out.append(f"**{d['tab']} — modified ({len(d['modified'])})**:")
            for entry in d["modified"][:20]:
                changed_cols = ", ".join(c[0] for c in entry["changes"][:5])
                if len(entry["changes"]) > 5:
                    changed_cols += f", … +{len(entry['changes']) - 5} more cell(s)"
                out.append(f"- `{entry['key']}` — {changed_cols}")
            if len(d["modified"]) > 20:
                out.append(f"- … {len(d['modified']) - 20} more")
            out.append("")

    return "\n".join(out).rstrip() + "\n"


def _maybe_truncate(markdown, max_bytes):
    """If markdown exceeds max_bytes, truncate with a marker."""
    encoded = markdown.encode("utf-8")
    if len(encoded) <= max_bytes:
        return markdown, False
    marker = (
        "\n\n---\n\n"
        "**Diff truncated for GitHub comment size cap.** "
        "Full diff is available as a sidecar file from whoever ran the tool.\n"
    )
    budget = max_bytes - len(marker.encode("utf-8"))
    return markdown.encode("utf-8")[:budget].decode("utf-8", errors="ignore") + marker, True


def main():
    parser = argparse.ArgumentParser(description="Diff two DD-reference XLSX workbooks.")
    parser.add_argument("baseline", help="Path to the baseline (current golden master) XLSX")
    parser.add_argument("new", help="Path to the new (incoming) XLSX")
    parser.add_argument("--json", action="store_true", help="Emit JSON instead of markdown")
    parser.add_argument(
        "--changelog",
        action="store_true",
        help="Emit a compact CHANGELOG.md entry instead of the detailed diff.",
    )
    parser.add_argument(
        "--ticket",
        type=str,
        default=None,
        help="Transport ticket number to link from the changelog entry (e.g. 203).",
    )
    parser.add_argument(
        "--title",
        type=str,
        default=None,
        help="Title for the changelog entry (e.g. 'DD 2.1 Sheet Update: ...').",
    )
    parser.add_argument(
        "--date",
        type=str,
        default=None,
        help="ISO date to stamp the changelog entry (default: today's date).",
    )
    parser.add_argument(
        "--max-bytes",
        type=int,
        default=60000,
        help="Max byte size for markdown output (default 60000; GitHub comment cap is 65,536).",
    )
    args = parser.parse_args()

    baseline_path = Path(args.baseline)
    new_path = Path(args.new)
    if not baseline_path.exists():
        sys.stderr.write(f"error: baseline not found: {baseline_path}\n")
        sys.exit(2)
    if not new_path.exists():
        sys.stderr.write(f"error: new not found: {new_path}\n")
        sys.exit(2)

    wb_baseline = load_workbook(baseline_path, data_only=True, read_only=False)
    wb_new = load_workbook(new_path, data_only=True, read_only=False)

    tab_diffs = []
    for tab_name, config in TAB_CONFIG.items():
        baseline_rows, _ = _load_tab(wb_baseline, tab_name, config)
        new_rows, _ = _load_tab(wb_new, tab_name, config)
        tab_diffs.append(_diff_tab(tab_name, baseline_rows, new_rows, config))

    if args.json:
        print(json.dumps({
            "baseline": str(baseline_path),
            "new": str(new_path),
            "tabs": tab_diffs,
        }, indent=2))
    elif args.changelog:
        # Compact entry for prepending to CHANGELOG.md
        from datetime import date as _date
        date_str = args.date or _date.today().isoformat()
        entry = _format_changelog(
            baseline_path, new_path, tab_diffs, args.ticket, args.title, date_str
        )
        print(entry)
    else:
        md = _format_markdown(baseline_path, new_path, tab_diffs)
        md, was_truncated = _maybe_truncate(md, args.max_bytes)
        print(md)
        if was_truncated:
            sys.stderr.write(
                f"note: markdown output truncated to {args.max_bytes:,} bytes for "
                f"GitHub comment cap. Re-run with a higher --max-bytes for the full diff.\n"
            )


if __name__ == "__main__":
    main()
