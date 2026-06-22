#!/usr/bin/env python3
"""
DD Sheet Linter — read-only validation of a RESO Data Dictionary XLSX sheet.

Port of the former dd-sheet-linter.js (openpyxl instead of SheetJS). This is the single read-only
validator for the DD sheets; lint-dd-sheet.py is its write-side counterpart (URL canonicalization +
whitespace normalization). The SimpleDataType/LookupStatus agreement check moved here from
lint-dd-sheet.py so all validation lives in one place.

Usage: python3 dd-sheet-linter.py <path-to-xlsx> [--prev <path-to-previous-xlsx>]

Checks:
  1. Sheet structure (expected sheets + columns)
  2. PascalCase on all StandardNames
  3. No duplicate (Resource, FieldName) pairs
  4. No duplicate (LookupName, StandardLookupValue) pairs
  5. Unicode cleanliness (no BOM, NBSP, ZWSP, ZWNJ, ZWJ)
  6. Referential integrity (LookupName -> field with lookup type)
  7. Deprecation tracking vs. previous version (warnings, not errors)
  8. Synonym not equal to any StandardName
  9. Version Info sheet has valid version
 10. No empty StandardName / StandardLookupValue cells in data rows
 11. SimpleDataType and LookupStatus agree (a field is an enumeration IFF its type is a String List)

Exits non-zero if any errors are found.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook

PASCAL_RE = re.compile(r"^[A-Z][a-zA-Z0-9]*$")
# RESO DD enumeration data types — only these legitimately carry a LookupStatus.
ENUM_DATA_TYPES = {"String List, Single", "String List, Multi"}
ZERO_WIDTH = {
    "﻿": "BOM (U+FEFF)",
    " ": "NBSP (U+00A0)",
    "​": "ZWSP (U+200B)",
    "‌": "ZWNJ (U+200C)",
    "‍": "ZWJ (U+200D)",
}


def sheet_to_dicts(ws) -> list[dict]:
    """Row 1 is the header; each data row becomes a dict that omits empty cells."""
    rows = ws.iter_rows()
    header = [cell.value for cell in next(rows)]
    out = []
    for row in rows:
        record = {}
        for key, cell in zip(header, row):
            if key is not None and cell.value is not None:
                record[key] = cell.value
        out.append(record)
    return out


def type_status_warnings(fields_data, field_name_col="StandardName", resource_col="ResourceName") -> list[str]:
    """SimpleDataType and LookupStatus must agree: a field is an enumeration IF AND ONLY IF its data
    type is 'String List, Single/Multi'. Flags both directions of disagreement —
      (a) a LookupStatus on a non-enumeration type — e.g. BuiltPre1978YN (a Boolean carrying 'Open',
          fixed by clearing the status) or HistoryTransactional.ClassName (a String that should be
          'String List, Single', fixed by correcting the type); and
      (b) an enumeration data type carrying no LookupStatus.
    Warnings are qualified by Resource.Field because a field name (e.g. ClassName) can appear on many
    resources with differing rows. Returns the warning messages; the caller decides severity."""
    out = []
    for row in fields_data:
        field_name = row.get(field_name_col)
        if not field_name:
            continue
        data_type = row.get("SimpleDataType")
        lookup_status = row.get("LookupStatus")
        resource = row.get(resource_col)
        label = f"{resource}.{field_name}" if resource else str(field_name)
        is_enum_type = str(data_type) in ENUM_DATA_TYPES
        has_status = lookup_status not in (None, "")
        if has_status and not is_enum_type:
            out.append(
                f"{label}: SimpleDataType={data_type!r} carries LookupStatus={lookup_status!r} "
                f"but is not an enumeration data type"
            )
        elif is_enum_type and not has_status:
            out.append(
                f"{label}: SimpleDataType={data_type!r} is an enumeration data type "
                f"but carries no LookupStatus"
            )
    return out


def main() -> None:
    args = sys.argv[1:]
    if not args:
        print("Usage: dd-sheet-linter.py <path-to-xlsx> [--prev <path-to-previous-xlsx>]")
        sys.exit(0)

    sheet_path = args[0]
    prev_path = args[args.index("--prev") + 1] if "--prev" in args else None

    errors: list[tuple[str, str]] = []
    warnings: list[tuple[str, str]] = []
    info: list[tuple[str, str]] = []
    add_error = lambda check, msg: errors.append((check, msg))
    add_warning = lambda check, msg: warnings.append((check, msg))
    add_info = lambda check, msg: info.append((check, msg))

    wb = load_workbook(Path(sheet_path).resolve(), read_only=True, data_only=True)
    prev_wb = load_workbook(Path(prev_path).resolve(), read_only=True, data_only=True) if prev_path else None

    # ── 1. Sheet structure ──
    EXPECTED_MINIMAL = ["Fields", "Lookups"]
    EXPECTED_FULL = ["Fields", "Lookups", "Changes", "Summary of Changes", "Version Info"]
    for name in EXPECTED_MINIMAL:
        if name not in wb.sheetnames:
            add_error("structure", f"Missing required sheet: {name}")
    for name in EXPECTED_FULL:
        if name not in wb.sheetnames:
            add_warning("structure", f"Missing optional sheet: {name}")

    if "Fields" not in wb.sheetnames or "Lookups" not in wb.sheetnames:
        print("Cannot proceed without Fields and Lookups sheets.")
        sys.exit(1)

    fields_data = sheet_to_dicts(wb["Fields"])
    lookups_data = sheet_to_dicts(wb["Lookups"])

    add_info("counts", f"Fields: {len(fields_data)} rows")
    add_info("counts", f"Lookups: {len(lookups_data)} rows")

    # ── Column name detection ──
    # Column presence comes from the union of all rows: sheet_to_dicts omits empty cells, so a blank
    # value in the first data row must not hide a column (e.g. an empty LookupStatus on row 1 would
    # otherwise gate off the agreement check).
    field_headers = {key for row in fields_data for key in row}
    lookup_headers = {key for row in lookups_data for key in row}

    def pick(headers, *candidates, default=None):
        for c in candidates:
            if c in headers:
                return c
        return default

    field_name_col = pick(field_headers, "StandardName", "FieldName", default="StandardName")
    resource_col = pick(field_headers, "ResourceName", "Resource", default="ResourceName")
    lookup_name_col = pick(lookup_headers, "LookupName", default="LookupName")
    lookup_value_col = pick(lookup_headers, "StandardLookupValue", "LookupDisplayName", default="StandardLookupValue")
    synonym_col = pick(field_headers, "Synonyms")

    # ── 2. PascalCase ──
    all_field_names = set()
    for row in fields_data:
        name = row.get(field_name_col)
        if not name:
            continue
        all_field_names.add(name)
        if not PASCAL_RE.match(str(name)):
            add_error("pascal-case", f'Field "{name}" is not PascalCase')

    # ── 3. Duplicate fields ──
    field_pairs = set()
    for row in fields_data:
        resource = row.get(resource_col)
        name = row.get(field_name_col)
        if not resource or not name:
            continue
        key = f"{resource}.{name}"
        if key in field_pairs:
            add_error("duplicate-field", f"Duplicate field: {key}")
        field_pairs.add(key)

    # ── 4. Duplicate lookups ──
    lookup_pairs = set()
    for row in lookups_data:
        name = row.get(lookup_name_col)
        value = row.get(lookup_value_col)
        if not name or not value:
            continue
        key = f"{name}.{value}"
        if key in lookup_pairs:
            add_error("duplicate-lookup", f"Duplicate lookup: {key}")
        lookup_pairs.add(key)

    # ── 5. Unicode cleanliness ──
    def check_unicode(ws, sheet_name):
        for r, row in enumerate(ws.iter_rows(), start=1):
            for c, cell in enumerate(row, start=1):
                val = cell.value
                if not isinstance(val, str):
                    continue
                for ch, label in ZERO_WIDTH.items():
                    if ch in val:
                        add_warning("unicode", f"{label} in {sheet_name} row {r}, col {c}")

    check_unicode(wb["Fields"], "Fields")
    check_unicode(wb["Lookups"], "Lookups")

    # ── 6. Referential integrity (LookupName -> field) ──
    lookup_field_names = {row.get(lookup_name_col) for row in lookups_data if row.get(lookup_name_col)}
    for lookup_name in lookup_field_names:
        name = str(lookup_name)
        if name not in all_field_names and "Type" not in name and "Status" not in name:
            add_info("referential", f'LookupName "{name}" has no exact field match (may be shared)')

    # ── 7. Deprecation tracking vs. previous version ──
    if prev_wb is not None and "Fields" in prev_wb.sheetnames:
        prev_fields = sheet_to_dicts(prev_wb["Fields"])
        prev_field_names = {row.get(field_name_col) for row in prev_fields if row.get(field_name_col)}
        removed = [n for n in prev_field_names if n not in all_field_names]
        added = [n for n in all_field_names if n not in prev_field_names]
        if removed:
            shown = ", ".join(str(n) for n in removed[:10]) + ("..." if len(removed) > 10 else "")
            add_warning("deprecation", f"{len(removed)} field(s) removed (deprecated): {shown}")
        if added:
            shown = ", ".join(str(n) for n in added[:10]) + ("..." if len(added) > 10 else "")
            add_info("additions", f"{len(added)} field(s) added: {shown}")

        if "Lookups" in prev_wb.sheetnames:
            prev_lookups = sheet_to_dicts(prev_wb["Lookups"])
            prev_keys = {
                f"{r.get(lookup_name_col)}.{r.get(lookup_value_col)}"
                for r in prev_lookups
                if r.get(lookup_name_col) and r.get(lookup_value_col)
            }
            curr_keys = {
                f"{r.get(lookup_name_col)}.{r.get(lookup_value_col)}"
                for r in lookups_data
                if r.get(lookup_name_col) and r.get(lookup_value_col)
            }
            removed_l = [k for k in prev_keys if k not in curr_keys]
            added_l = [k for k in curr_keys if k not in prev_keys]
            if removed_l:
                shown = ", ".join(removed_l[:5]) + ("..." if len(removed_l) > 5 else "")
                add_warning("deprecation", f"{len(removed_l)} lookup(s) removed: {shown}")
            if added_l:
                shown = ", ".join(added_l[:5]) + ("..." if len(added_l) > 5 else "")
                add_info("additions", f"{len(added_l)} lookup(s) added: {shown}")

    # ── 8. Synonym not equal to any StandardName ──
    if synonym_col:
        for row in fields_data:
            synonyms = row.get(synonym_col)
            if not isinstance(synonyms, str):
                continue
            for syn in (s.strip() for s in synonyms.split(",")):
                if syn and syn in all_field_names:
                    add_warning(
                        "synonym-collision",
                        f'Synonym "{syn}" for field "{row.get(field_name_col)}" is also a StandardName',
                    )

    # ── 9. Version Info ──
    if "Version Info" in wb.sheetnames:
        if not sheet_to_dicts(wb["Version Info"]):
            add_warning("version", "Version Info sheet is empty")

    # ── 10. No empty StandardName / StandardLookupValue ──
    empty_name = sum(1 for row in fields_data if not str(row.get(field_name_col) or "").strip())
    if empty_name:
        add_error("empty-name", f"{empty_name} field row(s) with empty StandardName")
    empty_lookup = sum(1 for row in lookups_data if not str(row.get(lookup_value_col) or "").strip())
    if empty_lookup:
        add_error("empty-lookup", f"{empty_lookup} lookup row(s) with empty StandardLookupValue")

    # ── 11. SimpleDataType / LookupStatus agreement ──
    if "SimpleDataType" in field_headers and "LookupStatus" in field_headers:
        for msg in type_status_warnings(fields_data, field_name_col, resource_col):
            add_warning("type-status-agreement", msg)

    # ── Report ──
    print("\n╔══════════════════════════════════════════╗")
    print("║         DD Sheet Linter Report           ║")
    print("╚══════════════════════════════════════════╝\n")
    print(f"File: {sheet_path}")
    if prev_path:
        print(f"Prev: {prev_path}")
    print()

    if info:
        print("ℹ  Info:")
        for check, msg in info:
            print(f"   [{check}] {msg}")
        print()
    if warnings:
        print(f"⚠  Warnings ({len(warnings)}):")
        for check, msg in warnings:
            print(f"   [{check}] {msg}")
        print()
    if errors:
        print(f"✗  Errors ({len(errors)}):")
        for check, msg in errors:
            print(f"   [{check}] {msg}")
        print()

    if not errors and not warnings:
        print("✓  All checks passed.\n")
    elif not errors:
        print(f"✓  No errors. {len(warnings)} warning(s).\n")
    else:
        print(f"✗  {len(errors)} error(s), {len(warnings)} warning(s).\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
