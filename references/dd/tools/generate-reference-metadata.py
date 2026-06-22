#!/usr/bin/env python3
"""
Generate reference metadata JSON from DD XLSX sheets.

Produces the same JSON structure as the Commander's MetadataReport serializer, extended with:
  - models[]       — resource/complex type definitions (ready for DD 2.2)
  - isEnumeration  — computed from SimpleDataType/LookupStatus
  - actions[] / functions[] — placeholders for future OData support
  - sourceResource on navigation properties

Port of the former generate-reference-metadata.js (openpyxl instead of SheetJS), byte-for-byte
output-compatible. Uses openpyxl so the DD tooling is single-runtime; the only behavioural subtlety
is matching JavaScript's String()/Number() coercion and JSON.stringify formatting (see js_string /
to_number / the json.dumps call).

Usage: python3 generate-reference-metadata.py <xlsx-path> <version> [output-path]
"""

from __future__ import annotations

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

# ── Type mapping ──

SIMPLE_TYPE_TO_EDM = {
    "Boolean": "Edm.Boolean",
    "Date": "Edm.Date",
    "Decimal": "Edm.Decimal",
    "Integer": "Edm.Int64",
    "Number": "Edm.Decimal",
    "String": "Edm.String",
    "String List, Single": "Edm.String",
    "String List, Multi": "Edm.String",
    "Timestamp": "Edm.DateTimeOffset",
    "Resource": "Edm.NavigationProperty",
}

ENUM_TYPES = {"String List, Single", "String List, Multi"}


def js_string(v) -> str:
    """Mirror JavaScript String(v): integral floats render without a trailing '.0', booleans as
    lowercase. openpyxl can hand back 50.0 where SheetJS produced 50; this keeps the JSON identical."""
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def to_number(v):
    """Mirror JavaScript Number(v) for the numeric facet columns: an int when integral, else a
    float; None when not a finite number (the JS guard is `x && !isNaN(Number(x))`)."""
    try:
        n = float(v)
    except (TypeError, ValueError):
        return None
    if n != n:  # NaN
        return None
    return int(n) if n.is_integer() else n


def sheet_to_dicts(ws) -> list[dict]:
    """Equivalent of SheetJS sheet_to_json: row 1 is the header; each data row becomes a dict that
    omits empty cells (so a missing key reads as None, matching `undefined` in the JS guards)."""
    rows = ws.iter_rows()
    header = [cell.value for cell in next(rows)]
    out = []
    for row in rows:
        record = {}
        for key, cell in zip(header, row):
            if key is not None and cell.value is not None:
                record[key] = cell.value
        out.append(record)
    return out


def build_fields(fields_raw: list[dict]) -> tuple[list[dict], set]:
    resource_set: set = set()
    fields: list[dict] = []

    for row in fields_raw:
        resource_name = row.get("ResourceName")
        field_name = row.get("StandardName")
        if not resource_name or not field_name:
            continue

        resource_set.add(js_string(resource_name))
        field_name_s = js_string(field_name)

        raw_simple = row.get("SimpleDataType")
        simple_type = js_string(raw_simple) if raw_simple is not None else ""
        sr_raw = row.get("SourceResource")
        source_resource = js_string(sr_raw) if sr_raw else None
        ln_raw = row.get("LookupName")
        lookup_name = js_string(ln_raw) if ln_raw else None
        sug_max_precision = row.get("SugMaxPrecision")
        max_length = row.get("SugMaxLength")
        ls_raw = row.get("LookupStatus")

        is_expansion = simple_type in ("Resource", "Collection") or bool(source_resource)
        is_collection = simple_type in ("String List, Multi", "Collection")
        is_enumeration = simple_type in ENUM_TYPES or (
            bool(ls_raw) and js_string(ls_raw).strip() != ""
        )
        target_resource = (source_resource or field_name_s) if is_expansion else None
        # Only EntityEventSequence needs Int64; other Number fields stay Decimal.
        resolved_simple_type = (
            "Integer" if simple_type == "Number" and field_name_s == "EntityEventSequence"
            else simple_type
        )
        if is_expansion:
            edm_type = (
                f"Collection(org.reso.metadata.{target_resource})" if is_collection
                else f"org.reso.metadata.{target_resource}"
            )
        elif is_enumeration and lookup_name:
            edm_type = f"org.reso.metadata.enums.{lookup_name}"
        else:
            edm_type = SIMPLE_TYPE_TO_EDM.get(resolved_simple_type, resolved_simple_type)

        annotations = []
        display_name = row.get("DisplayName")
        if display_name:
            annotations.append({"term": "RESO.OData.Metadata.StandardName", "value": js_string(display_name)})
        wiki_url = row.get("WikiPageUrl")
        if wiki_url:
            annotations.append({"term": "RESO.DDWikiUrl", "value": js_string(wiki_url)})
        definition = row.get("Definition")
        if definition:
            annotations.append({"term": "Core.Description", "value": js_string(definition)})

        field = {
            "resourceName": js_string(resource_name),
            "fieldName": field_name_s,
            "type": edm_type,
            "nullable": True,
            "isEnumeration": is_enumeration,
            "annotations": annotations,
        }

        ml = to_number(max_length) if max_length else None
        if ml is not None:
            field["maxLength"] = ml

        # For decimals: DD SugMaxLength = OData precision (total digits),
        # DD SugMaxPrecision = OData scale (decimal places).
        if resolved_simple_type == "Number" and ml is not None:
            field["precision"] = ml
            smp = to_number(sug_max_precision) if sug_max_precision else None
            field["scale"] = smp if smp is not None else 0
        elif resolved_simple_type == "Decimal" and ml is not None:
            field["precision"] = ml
            smp = to_number(sug_max_precision) if sug_max_precision else None
            field["scale"] = smp if smp is not None else 2

        if is_collection:
            field["isCollection"] = True
        if is_expansion:
            field["isExpansion"] = True
            field["typeName"] = target_resource
            if source_resource:
                field["sourceResource"] = source_resource

        # DD metadata (useful for cert testing, not in Commander output).
        if ls_raw:
            field["lookupStatus"] = js_string(ls_raw)
        element_status = row.get("ElementStatus")
        if element_status:
            field["elementStatus"] = js_string(element_status)
        payloads = row.get("Payloads")
        if payloads:
            field["payloads"] = js_string(payloads)
        synonyms = row.get("Synonyms")
        if synonyms:
            field["synonyms"] = js_string(synonyms)

        fields.append(field)

    return fields, resource_set


def build_lookups(lookups_raw: list[dict]) -> list[dict]:
    lookups: list[dict] = []
    for row in lookups_raw:
        lookup_name = row.get("LookupName")
        slv = row.get("StandardLookupValue")
        standard_lookup_value = slv if slv is not None else row.get("LookupDisplayName")
        lov = row.get("LegacyODataValue")
        lookup_value = lov if lov is not None else standard_lookup_value
        if not lookup_name or not lookup_value:
            continue

        annotations = [
            {"term": "RESO.OData.Metadata.StandardName", "value": js_string(standard_lookup_value)}
        ]
        wiki_url = row.get("WikiPageUrl")
        if wiki_url:
            annotations.append({"term": "RESO.DDWikiUrl", "value": js_string(wiki_url)})
        definition = row.get("Definition")
        if definition:
            annotations.append({"term": "Core.Description", "value": js_string(definition)})
        legacy_value = row.get("LegacyODataValue")
        if legacy_value:
            annotations.append({"term": "RESO.OData.Metadata.LegacyODataValue", "value": js_string(legacy_value)})
        synonyms = row.get("Synonyms")
        if synonyms:
            annotations.append({"term": "RESO.OData.Metadata.Synonyms", "value": js_string(synonyms)})

        lookups.append({
            "lookupName": f"org.reso.metadata.enums.{js_string(lookup_name)}",
            "lookupValue": js_string(lookup_value),
            "type": "Edm.Int32",
            "annotations": annotations,
        })
    return lookups


def iso_now() -> str:
    """generatedOn timestamp. Prefer DD_GENERATED_ON from the environment so every file produced in a
    single build shares one timestamp (the workflow sets it once); fall back to the current time for
    ad-hoc local runs (millisecond precision + 'Z', matching JS new Date().toISOString())."""
    override = os.environ.get("DD_GENERATED_ON")
    if override:
        return override
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"


def main() -> None:
    args = sys.argv[1:]
    if len(args) < 2:
        print("Usage: generate-reference-metadata.py <xlsx-path> <version> [output-path]")
        sys.exit(0)

    xlsx_path, version = args[0], args[1]
    output_path = args[2] if len(args) >= 3 else None

    wb = load_workbook(Path(xlsx_path).resolve(), read_only=True, data_only=True)

    if "Fields" not in wb.sheetnames:
        print('No "Fields" sheet found.', file=sys.stderr)
        sys.exit(1)
    fields, resource_set = build_fields(sheet_to_dicts(wb["Fields"]))

    if "Lookups" not in wb.sheetnames:
        print('No "Lookups" sheet found.', file=sys.stderr)
        sys.exit(1)
    lookups = build_lookups(sheet_to_dicts(wb["Lookups"]))

    resources = sorted(resource_set)
    models = [{"modelName": name, "modelType": "EntityType"} for name in resources]

    output = {
        "description": f"RESO Data Dictionary {version} Reference Metadata",
        "version": version,
        "generatedOn": iso_now(),
        "resources": resources,
        "models": models,
        "fields": fields,
        "lookups": lookups,
        # Placeholders for future OData structural elements.
        "actions": [],
        "functions": [],
    }

    out_file = (
        Path(output_path).resolve() if output_path
        else Path(f"reso-certification/etl/reference-metadata/dd-{version}.json").resolve()
    )
    out_file.write_text(
        json.dumps(output, indent=2, ensure_ascii=False, separators=(",", ": ")),
        encoding="utf-8",
    )

    enum_count = sum(1 for f in fields if f["isEnumeration"])
    expansion_count = sum(1 for f in fields if f.get("isExpansion"))
    print(f"Generated: {out_file}")
    print(f"  Version: {version}")
    print(f"  Resources: {len(resource_set)}")
    print(f"  Models: {len(models)} (all EntityType for now)")
    print(f"  Fields: {len(fields)} ({enum_count} enumerations, {expansion_count} expansions)")
    print(f"  Lookups: {len(lookups)}")
    print("  Actions: 0 (placeholder)")
    print("  Functions: 0 (placeholder)")


if __name__ == "__main__":
    main()
