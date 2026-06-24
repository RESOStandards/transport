"""
Tests for dd-sheet-linter.py — the DD XLSX read-only validator.

The validator is cert-adjacent tooling, so its checks must hold 100%. Run with the DD-lint venv
(openpyxl + pytest), see the "DD reference tooling" section of CLAUDE.md:

    references/dd/tools/.venv-dd-lint/bin/python -m pytest references/dd/tools/test_dd_sheet_linter.py

The venv is intentionally kept out of git.
"""

from __future__ import annotations

import importlib.util
import os
import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook

# Load the hyphenated script by path.
_LINTER_PATH = Path(__file__).parent / "dd-sheet-linter.py"
_spec = importlib.util.spec_from_file_location("dd_sheet_linter", _LINTER_PATH)
assert _spec and _spec.loader
linter = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(linter)


def _fields(*rows) -> list[dict]:
    """Build the list-of-dicts shape sheet_to_dicts produces — empty (None) cells omitted, so a
    missing key reads as absent."""
    cols = ["ResourceName", "StandardName", "SimpleDataType", "LookupStatus"]
    return [{c: v for c, v in zip(cols, r) if v is not None} for r in rows]


# --- sheet_to_dicts ------------------------------------------------------------------------------


def test_sheet_to_dicts_omits_empty_cells():
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B", "C"])
    ws.append([1, None, 3])
    ws.append([None, None, None])
    rows = linter.sheet_to_dicts(ws)
    assert rows[0] == {"A": 1, "C": 3}  # B omitted (empty)
    assert rows[1] == {}  # all-empty row


# --- type_status_warnings (SimpleDataType / LookupStatus agreement) ------------------------------


def test_agreement_passes_consistent_rows():
    rows = _fields(
        ("Property", "StandardStatus", "String List, Single", "Locked with Enumerations"),
        ("Property", "Appliances", "String List, Multi", "Open with Enumerations"),
        ("Property", "ListPrice", "Number", None),
        ("Property", "PublicRemarks", "String", None),
        ("Media", "MediaModificationTimestamp", "Timestamp", None),
    )
    assert linter.type_status_warnings(rows) == []


def test_agreement_flags_lookupstatus_on_non_enum():
    # Direction (a): the two real DD 2.1 cases.
    rows = _fields(
        ("HistoryTransactional", "ClassName", "String", "Open with Enumerations"),
        ("Property", "BuiltPre1978YN", "Boolean", "Open"),
    )
    warnings = linter.type_status_warnings(rows)
    assert len(warnings) == 2
    assert any(w.startswith("HistoryTransactional.ClassName:") and "not an enumeration" in w for w in warnings)
    assert any(w.startswith("Property.BuiltPre1978YN:") and "not an enumeration" in w for w in warnings)


def test_agreement_flags_enum_type_without_status():
    # Direction (b): an enumeration data type carrying no LookupStatus.
    rows = _fields(
        ("Property", "OrphanEnum", "String List, Single", None),
        ("Property", "OrphanMulti", "String List, Multi", ""),  # empty string counts as absent
    )
    warnings = linter.type_status_warnings(rows)
    assert len(warnings) == 2
    assert all("carries no LookupStatus" in w for w in warnings)


def test_agreement_qualifies_label_by_resource():
    # ClassName appears on many resources; only one row disagrees — the label must pinpoint it.
    rows = _fields(
        ("ContactListings", "ClassName", "String List, Single", "Open with Enumerations"),
        ("HistoryTransactional", "ClassName", "String", "Open with Enumerations"),
    )
    warnings = linter.type_status_warnings(rows)
    assert len(warnings) == 1
    assert warnings[0].startswith("HistoryTransactional.ClassName:")


def test_agreement_skips_rows_without_field_name():
    rows = _fields(("Property", None, "String", "Open"))  # StandardName missing -> skipped
    assert linter.type_status_warnings(rows) == []


# --- end-to-end (full validator over a fixture workbook) -----------------------------------------


def test_end_to_end_flags_and_exits_nonzero():
    wb = Workbook()
    fields = wb.active
    fields.title = "Fields"
    fields.append(["ResourceName", "StandardName", "SimpleDataType", "LookupStatus"])
    fields.append(["Property", "ListPrice", "Number", None])                  # clean
    fields.append(["Property", "badName", "String", None])                    # not PascalCase -> error
    fields.append(["HistoryTransactional", "ClassName", "String", "Open with Enumerations"])  # agreement warn
    lookups = wb.create_sheet("Lookups")
    lookups.append(["LookupName", "StandardLookupValue"])
    lookups.append(["Sewer", "Public Sewer"])

    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "fixture.xlsx")
        wb.save(path)
        result = subprocess.run(
            [sys.executable, str(_LINTER_PATH), path], capture_output=True, text=True
        )

    assert result.returncode == 1  # the PascalCase error fails the run
    assert "pascal-case" in result.stdout
    assert "badName" in result.stdout
    assert "type-status-agreement" in result.stdout
    assert "HistoryTransactional.ClassName" in result.stdout
